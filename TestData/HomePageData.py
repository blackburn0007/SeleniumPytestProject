import openpyxl


class HomePageData:
    test_HomePageData = [{"firstname": "John", "lastname": "Wick", "email": "john@example.com"},
                         {"firstname": "Smith", "lastname": "Wilson", "email": "smith@example.com"}]

    # Optional if we need to drive Data from Excel
    @staticmethod
    def getTestData(test_case_name):
        dictionary = {}
        book = openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dictionary]
